#!/usr/bin/env python3
"""
Research Paper Organization Tool
---------------------------------
Scans a folder of research papers (PDFs), extracts metadata, and generates
a literature review spreadsheet following a specific template format.

Handles Dropbox files that may not be fully downloaded locally.
"""

import os
import sys
import argparse
from pathlib import Path
from datetime import datetime
import re
import time

# PDF metadata extraction
try:
    import PyPDF2
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    print("Warning: PyPDF2 not installed. Install with: pip install PyPDF2")

# HTTP requests for API calls
try:
    import requests
    REQUESTS_SUPPORT = True
except ImportError:
    REQUESTS_SUPPORT = False
    print("Warning: requests not installed. Install with: pip install requests")

# Excel writing
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("Warning: openpyxl not installed. Install with: pip install openpyxl")


class PaperMetadata:
    """Store metadata extracted from a research paper."""
    
    def __init__(self, filepath):
        self.filepath = Path(filepath)
        self.filename = self.filepath.name
        self.folder = self.filepath.parent.name
        self.relative_path = None
        
        # Metadata fields
        self.title = None
        self.author = None
        self.year = None
        self.concept = None  # Will be derived from folder structure
        self.journal = None
        self.url = None
        
        # Status
        self.is_downloaded = self.check_if_downloaded()
        self.file_size = self.get_file_size()
        
    def check_if_downloaded(self):
        """Check if Dropbox file is actually downloaded."""
        if not self.filepath.exists():
            return False
        
        # Dropbox placeholder files are typically very small
        # and may have specific attributes
        try:
            size = self.filepath.stat().st_size
            # If file is less than 1KB, it might be a placeholder
            if size < 1024:
                return False
            return True
        except:
            return False
    
    def get_file_size(self):
        """Get file size in MB."""
        try:
            size_bytes = self.filepath.stat().st_size
            return size_bytes / (1024 * 1024)
        except:
            return 0
    
    def extract_metadata_from_pdf(self):
        """Extract metadata from PDF file. Tries DOI lookup first, then falls back to PDF parsing."""
        if not PDF_SUPPORT or not self.is_downloaded:
            return
        
        try:
            with open(self.filepath, 'rb') as f:
                pdf = PyPDF2.PdfReader(f)
                
                # First, try to extract DOI and fetch metadata from Crossref API
                doi = None
                if len(pdf.pages) > 0:
                    # Extract DOI from first few pages
                    for page_num in range(min(3, len(pdf.pages))):
                        page_text = pdf.pages[page_num].extract_text()
                        doi = self._extract_doi_from_text(page_text)
                        if doi:
                            break
                
                # If DOI found, try to fetch metadata from Crossref
                if doi and REQUESTS_SUPPORT:
                    if self._fetch_metadata_from_crossref(doi):
                        print(f"  ✓ Metadata fetched from Crossref (DOI: {doi})")
                        return  # Successfully got metadata from API
                
                # Fall back to PDF metadata and text extraction
                # Get PDF metadata
                if pdf.metadata:
                    self.title = pdf.metadata.get('/Title', None)
                    self.author = pdf.metadata.get('/Author', None)
                    
                    # Try to extract year from creation date
                    creation_date = pdf.metadata.get('/CreationDate', '')
                    year_match = re.search(r'(\d{4})', str(creation_date))
                    if year_match:
                        self.year = year_match.group(1)
                
                # Extract from first page text if metadata is missing
                if len(pdf.pages) > 0:
                    first_page = pdf.pages[0].extract_text()
                    lines = [line.strip() for line in first_page.split('\n') if line.strip()]
                    
                    # Extract title from first page
                    if not self.title:
                        self.title = self._extract_title_from_text(lines)
                    
                    # If title not found on first page, try second page (some PDFs have cover pages)
                    if not self.title and len(pdf.pages) > 1:
                        second_page = pdf.pages[1].extract_text()
                        lines2 = [line.strip() for line in second_page.split('\n') if line.strip()]
                        self.title = self._extract_title_from_text(lines2)
                    
                    # Extract authors from first page
                    if not self.author:
                        self.author = self._extract_authors_from_text(lines)
                    
                    # If authors not found on first page, try second page
                    if not self.author and len(pdf.pages) > 1:
                        second_page = pdf.pages[1].extract_text()
                        lines2 = [line.strip() for line in second_page.split('\n') if line.strip()]
                        self.author = self._extract_authors_from_text(lines2)
                    
                    # Extract year from text if not found
                    if not self.year:
                        self.year = self._extract_year_from_text(first_page)
        
        except Exception as e:
            print(f"  Error extracting PDF metadata from {self.filename}: {e}")
    
    def _extract_doi_from_text(self, text):
        """Extract DOI from text using regex patterns."""
        if not text:
            return None
        
        # Common DOI patterns:
        # "DOI: 10.1234/abc.123"
        # "doi:10.1234/abc.123"
        # "https://doi.org/10.1234/abc.123"
        # "10.1234/abc.123"
        
        patterns = [
            r'doi\.org/(10\.\d+/[^\s\)]+)',  # https://doi.org/10.1234/abc
            r'DOI[:\s]+(10\.\d+/[^\s\)]+)',  # DOI: 10.1234/abc
            r'\b(10\.\d+/[^\s\)]+)',  # Just the DOI number
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                doi = match.group(1).strip()
                # Clean up trailing punctuation
                doi = re.sub(r'[.,;:]+$', '', doi)
                return doi
        
        return None
    
    def _fetch_metadata_from_crossref(self, doi):
        """Fetch metadata from Crossref API using DOI."""
        if not REQUESTS_SUPPORT:
            return False
        
        try:
            # Crossref API endpoint (free, no API key needed)
            url = f"https://api.crossref.org/works/{doi}"
            headers = {
                'User-Agent': 'Research-Paper-Organizer/1.0 (mailto:user@example.com)'  # Polite API usage
            }
            
            response = requests.get(url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                if 'message' in data:
                    msg = data['message']
                    
                    # Extract title
                    if 'title' in msg and msg['title']:
                        titles = msg['title']
                        if isinstance(titles, list) and len(titles) > 0:
                            self.title = titles[0]
                        elif isinstance(titles, str):
                            self.title = titles
                    
                    # Extract authors
                    if 'author' in msg and msg['author']:
                        authors = []
                        for author in msg['author']:
                            # Format: "LastName, FirstName" or "FirstName LastName"
                            given = author.get('given', '')
                            family = author.get('family', '')
                            if family:
                                if given:
                                    authors.append(f"{given} {family}")
                                else:
                                    authors.append(family)
                        
                        if authors:
                            # Limit to first 6 authors
                            self.author = ', '.join(authors[:6])
                    
                    # Extract year
                    if 'published-print' in msg and msg['published-print']:
                        date_parts = msg['published-print'].get('date-parts', [])
                        if date_parts and len(date_parts[0]) > 0:
                            self.year = str(date_parts[0][0])
                    elif 'published-online' in msg and msg['published-online']:
                        date_parts = msg['published-online'].get('date-parts', [])
                        if date_parts and len(date_parts[0]) > 0:
                            self.year = str(date_parts[0][0])
                    elif 'issued' in msg and msg['issued']:
                        date_parts = msg['issued'].get('date-parts', [])
                        if date_parts and len(date_parts[0]) > 0:
                            self.year = str(date_parts[0][0])
                    
                    # Extract journal
                    if 'container-title' in msg and msg['container-title']:
                        containers = msg['container-title']
                        if isinstance(containers, list) and len(containers) > 0:
                            self.journal = containers[0]
                        elif isinstance(containers, str):
                            self.journal = containers
                    
                    # Extract URL
                    if 'URL' in msg:
                        self.url = msg['URL']
                    elif 'link' in msg and msg['link']:
                        # Get first link
                        links = msg['link']
                        if isinstance(links, list) and len(links) > 0:
                            self.url = links[0].get('URL', '')
                    
                    return True
            
            # Rate limiting - wait a bit if we get 429
            elif response.status_code == 429:
                print(f"  Rate limited, waiting 1 second...")
                time.sleep(1)
            
            return False
            
        except requests.exceptions.RequestException as e:
            # Silently fail - we'll fall back to PDF parsing
            return False
        except Exception as e:
            # Silently fail - we'll fall back to PDF parsing
            return False
    
    def _extract_title_from_text(self, lines):
        """Extract title from first page lines."""
        if not lines:
            return None
        
        # First, check if line 0 has arXiv info with title
        if len(lines) > 0 and 'arXiv:' in lines[0]:
            line = lines[0]
            # Try to extract title after arXiv info
            # Pattern: "arXiv:XXXX.XXXX [category] DateTitle" or "arXiv:XXXX.XXXX Title"
            # Remove arXiv ID
            after_arxiv = re.sub(r'^.*?arXiv:[^\s]+\s*', '', line)
            # Remove category in brackets
            after_arxiv = re.sub(r'^\[[^\]]+\]\s*', '', after_arxiv)
            # Remove date pattern (e.g., "9 Apr 2008")
            after_arxiv = re.sub(r'^\d{1,2}\s+\w+\s+\d{4}', '', after_arxiv)
            title = after_arxiv.strip()
            # Clean up any trailing artifacts
            title = re.sub(r'\s+', ' ', title)  # Normalize whitespace
            if 20 <= len(title) <= 200:
                return title
        
        # Check for multi-line titles FIRST (especially if first line is short)
        # Titles often span 2-3 lines, and we should check this before single-line matches
        for i in range(min(5, len(lines) - 1)):
            # Skip if first line looks like metadata
            if re.search(r'^\(|^Received|^Published|^DOI:|^www\.|^http', lines[i], re.IGNORECASE):
                continue
            
            # Skip if line looks like author line
            if ',' in lines[i] and re.search(r'\d+[,\s]', lines[i]) and len(lines[i]) < 150:
                continue
            
            # Try combining 2-3 consecutive lines starting from line i
            for num_lines in [2, 3]:
                if i + num_lines <= len(lines):
                    # Skip if any line in the combination looks like an author line
                    skip_combination = False
                    for j in range(i, i + num_lines):
                        if ',' in lines[j] and re.search(r'\d+[,\s]', lines[j]) and len(lines[j]) < 150:
                            skip_combination = True
                            break
                    if skip_combination:
                        continue
                    
                    combined = ' '.join(lines[i:i+num_lines])
                    # Check if combined looks like a title
                    if (30 <= len(combined) <= 250 and 
                        combined[0].isupper() and
                        '@' not in combined and
                        'http' not in combined.lower() and
                        not re.search(r'^DOI:|^www\.', combined, re.IGNORECASE) and
                        # Make sure it doesn't start with author-like patterns
                        not re.match(r'^[A-Z][a-z]+\s+[A-Z]', combined)):  # Not "FirstName LastName"
                        # Clean up and return
                        title = re.sub(r'\s+', ' ', combined).strip()
                        return title
        
        # If no multi-line title found, try single-line titles
        for i, line in enumerate(lines[:8]):
            # Skip lines that look like dates, received info, or are too short
            if re.search(r'^\(|^Received|^Published|^DOI:|^www\.|^http', line, re.IGNORECASE):
                continue
            
            # Skip lines that look like author lines (contain many commas and numbers)
            if ',' in line and re.search(r'\d+[,\s]', line):
                # This might be an author line, skip unless it's very long (unlikely for authors)
                if len(line) < 150:
                    continue
            
            # Skip if line looks like it starts with a name (FirstName LastName pattern)
            # But be careful - titles can also start with capitalized words
            # Only skip if it looks like a short name pattern (2-3 words, reasonable length)
            if re.match(r'^[A-Z][a-z]+\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+)?$', line) and len(line) < 50:
                continue
            
            # Title is usually a meaningful line that doesn't contain email addresses or URLs
            if (20 <= len(line) <= 200 and 
                '@' not in line and 
                'http' not in line.lower() and
                not re.match(r'^\d+[,\s]', line)):  # Doesn't start with numbers
                # Prefer first line if it looks like a title (capitalized)
                if i == 0 and line[0].isupper():
                    return line
                # Or use early lines that are capitalized
                elif i < 3 and line[0].isupper():
                    return line
        
        return None
    
    def _extract_authors_from_text(self, lines):
        """Extract author list from first page lines using regex patterns."""
        if not lines:
            return None
        
        # First, check for "et al." pattern - extract author name after "et al."
        for line in lines[:10]:
            # Look for "et al. Author Name" pattern (author comes AFTER et al.)
            et_al_match = re.search(r'et\s+al\.?\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*(?:\s+[A-Z]\.?[a-z]*)?)', line, re.IGNORECASE)
            if et_al_match:
                author_name = et_al_match.group(1).strip()
                # Make sure it's not part of a URL or website name, and not a common word
                if (not re.search(r'\.(org|com|edu|net)', line, re.IGNORECASE) and 
                    len(author_name) > 2 and
                    author_name.lower() not in ['science', 'nature', 'cell', 'journal', 'article', 'paper']):
                    return author_name
        
        # Try to extract multi-line author lists (authors often span 2-3 lines)
        # Start from line 1 (skip line 0 which is usually the title)
        for i in range(1, min(7, len(lines))):
            # Skip lines with URLs, websites, or metadata
            if re.search(r'http|www\.|\.org|\.com|\.edu|DOI|Received|Published|PACS|Abstract|Introduction', lines[i], re.IGNORECASE):
                continue
            
            # Skip if line looks like a title (no commas, capitalized, reasonable length)
            if (',' not in lines[i] and 
                len(lines[i]) > 20 and 
                len(lines[i]) < 200 and
                lines[i][0].isupper() and
                'http' not in lines[i].lower()):
                continue
            
            # Try combining 2-3 consecutive lines that might contain authors
            for num_lines in [2, 3]:
                if i + num_lines <= len(lines):
                    # Check if any line in range has URLs - skip if so
                    has_url = False
                    for j in range(i, i + num_lines):
                        if re.search(r'http|www\.|\.org|\.com|\.edu', lines[j], re.IGNORECASE):
                            has_url = True
                            break
                    if has_url:
                        continue
                    
                    # Combine lines, but skip if first line looks like a title
                    # (titles usually don't have commas and are capitalized)
                    first_line = lines[i]
                    if (',' not in first_line and 
                        len(first_line) > 15 and 
                        len(first_line) < 200 and
                        first_line[0].isupper() and
                        'http' not in first_line.lower()):
                        # This looks like a title, skip this combination
                        continue
                    
                    # Combine lines
                    combined = ' '.join(lines[i:i+num_lines])
                    authors = self._parse_author_line(combined)
                    if authors:
                        return authors
        
        # Look at individual lines for author information
        for line in lines[1:8]:
            # Skip lines that are clearly not author lines
            if re.search(r'^(Received|Published|DOI|PACS|Abstract|Introduction|Department|University|Institute|http|www\.|\.org|\.com|\.edu)', line, re.IGNORECASE):
                continue
            
            authors = self._parse_author_line(line)
            if authors:
                return authors
        
        return None
    
    def _parse_author_line(self, line):
        """Parse a line to extract author names."""
        if not line or len(line) < 5:
            return None
        
        # Skip if line contains URLs or website names
        if re.search(r'http|www\.|\.org|\.com|\.edu|sciencemag|arxiv', line, re.IGNORECASE):
            return None
        
        # Check if line looks like author list (contains commas or has name patterns)
        if ',' not in line and not re.search(r'[A-Z][a-z]+\s+[A-Z]', line):
            return None
        
        # Remove superscripts/affiliations (numbers, asterisks after names)
        # Pattern: "Name1,1,2" -> "Name1"
        author_line = re.sub(r',\s*\d+[,\s]*', ', ', line)
        author_line = re.sub(r'\*\s*', '', author_line)
        author_line = re.sub(r'\s+and\s+', ', ', author_line, flags=re.IGNORECASE)
        
        # Extract author names
        authors = []
        parts = [p.strip() for p in author_line.split(',')]
        
        for part in parts:
            part = part.strip()
            # Skip if it's just a number or very short
            if len(part) < 3 or part.isdigit():
                continue
            # Skip if it looks like an institution
            if any(word in part.lower() for word in ['university', 'department', 'institute', 'laboratory', 'center', 'centre', 'college', 'school']):
                continue
            # Skip if it's "et al" or contains URLs
            if 'et al' in part.lower() or re.search(r'\.(org|com|edu|net)', part, re.IGNORECASE):
                continue
            
            # Clean up the author name
            author = re.sub(r'^\d+[,\s]*', '', part)  # Remove leading numbers
            author = re.sub(r'[,\s]*\d+$', '', author)  # Remove trailing numbers
            author = author.strip()
            
            # Check if it looks like a name (has letters, reasonable length, capitalized)
            if (len(author) >= 3 and 
                re.search(r'[A-Za-z]', author) and
                # Should start with capital letter (name pattern)
                re.match(r'^[A-Z]', author)):
                authors.append(author)
        
        # If we found authors, join them
        if len(authors) >= 1:
            # Limit to first 6 authors to avoid noise
            author_list = ', '.join(authors[:6])
            if len(author_list) > 5:  # Make sure it's meaningful
                return author_list
        
        return None
    
    def _extract_year_from_text(self, text):
        """Extract year from text (look for publication dates, etc.)."""
        # Look for patterns like "(Received 14 March 2013; published 22 May 2013)"
        # or "(Dated: April 9, 2008)" or "2008" in context
        patterns = [
            r'\(.*?(\d{4}).*?\)',  # Year in parentheses
            r'\b(19|20)\d{2}\b',   # 4-digit year (1900-2099)
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, text[:2000])  # Check first 2000 chars
            if matches:
                # Get the last match (usually the publication year)
                if isinstance(matches[-1], tuple):
                    year = ''.join(matches[-1])
                else:
                    year = matches[-1]
                # Validate it's a reasonable year
                if 1900 <= int(year) <= 2100:
                    return year
        
        return None
    
    def extract_metadata_from_filename(self):
        """Try to extract year and author from filename."""
        # Common patterns: "Author_Year_Title.pdf" or "Author (Year) Title.pdf"
        filename_no_ext = self.filepath.stem
        
        # Look for year (4 digits)
        year_match = re.search(r'[(\[]?(\d{4})[)\]]?', filename_no_ext)
        if year_match and not self.year:
            self.year = year_match.group(1)
        
        # Try to extract author (usually first part before year or underscore)
        if not self.author:
            # Pattern: "LastName" or "LastName_FirstInitial" before year
            parts = re.split(r'[_\-\s]', filename_no_ext)
            if parts:
                potential_author = parts[0]
                # Clean up common prefixes
                potential_author = re.sub(r'^(the|a|an)\s+', '', potential_author, flags=re.IGNORECASE)
                if len(potential_author) > 2 and potential_author.isalpha():
                    self.author = potential_author
    
    def set_concept_from_folder(self, root_path):
        """Derive concept/topic from folder structure."""
        try:
            # Get relative path from root
            rel_path = self.filepath.relative_to(root_path)
            
            # Use parent folder name(s) as concept
            if len(rel_path.parents) > 1:
                # If nested, use all folder names except root
                folders = [p.name for p in rel_path.parents if p != Path('.')]
                folders.reverse()
                self.concept = ' > '.join(folders)
            else:
                self.concept = self.folder
            
            self.relative_path = str(rel_path)
        except:
            self.concept = "Uncategorized"


class PaperOrganizer:
    """Main class for organizing papers and generating spreadsheet."""
    
    def __init__(self, root_path, output_excel=None):
        self.root_path = Path(root_path)
        self.papers = []
        self.output_excel = output_excel or self.root_path / f"literature_review_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
    def scan_papers(self, extensions=['.pdf']):
        """Recursively scan for paper files."""
        print(f"Scanning folder: {self.root_path}")
        
        for ext in extensions:
            for filepath in self.root_path.rglob(f"*{ext}"):
                # Skip hidden files and system folders
                if any(part.startswith('.') for part in filepath.parts):
                    continue
                
                paper = PaperMetadata(filepath)
                paper.set_concept_from_folder(self.root_path)
                
                # Extract metadata
                if paper.is_downloaded:
                    print(f"  Processing: {paper.filename}")
                    paper.extract_metadata_from_pdf()
                    paper.extract_metadata_from_filename()
                else:
                    print(f"  Skipping (not downloaded): {paper.filename}")
                    paper.extract_metadata_from_filename()
                
                self.papers.append(paper)
        
        print(f"\nFound {len(self.papers)} papers")
        downloaded = sum(1 for p in self.papers if p.is_downloaded)
        print(f"  Downloaded: {downloaded}")
        print(f"  Not downloaded: {len(self.papers) - downloaded}")
    
    def generate_spreadsheet(self):
        """Generate Excel spreadsheet following the template format."""
        if not EXCEL_SUPPORT:
            print("Error: openpyxl not installed. Cannot generate Excel file.")
            return
        
        print(f"\nGenerating spreadsheet: {self.output_excel}")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Literature Review"
        
        # Define headers (matching template)
        headers = [
            'Concept', 'Author', 'Year', 'Title', 'Journal', 
            'Main idea', 'Conclusion', 'Notes 1', 'Notes 2', 
            'Cross-ref', 'Excerpts', 'URL'
        ]
        
        # Add filename and download status columns
        headers.extend(['Filename', 'Downloaded', 'File Path'])
        
        # Write headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sort papers by concept, then year
        sorted_papers = sorted(self.papers, key=lambda p: (
            p.concept or 'ZZZ',  # Put uncategorized at end
            p.year or '9999',
            p.author or 'ZZZ'
        ))
        
        # Write paper data
        current_concept = None
        row = 2
        
        for paper in sorted_papers:
            # Add a blank row between different concepts for readability
            if current_concept and paper.concept != current_concept:
                row += 1
            
            current_concept = paper.concept
            
            # Write data
            ws.cell(row=row, column=1, value=paper.concept)
            ws.cell(row=row, column=2, value=paper.author)
            ws.cell(row=row, column=3, value=paper.year)
            ws.cell(row=row, column=4, value=paper.title or paper.filename)
            ws.cell(row=row, column=5, value=paper.journal)
            # Columns 6-12 (Main idea, Conclusion, Notes, etc.) left empty for manual filling
            ws.cell(row=row, column=12, value=paper.url)
            
            # Additional columns
            ws.cell(row=row, column=13, value=paper.filename)
            ws.cell(row=row, column=14, value='Yes' if paper.is_downloaded else 'No')
            ws.cell(row=row, column=15, value=paper.relative_path)
            
            row += 1
        
        # Adjust column widths
        column_widths = {
            'A': 25,  # Concept
            'B': 20,  # Author
            'C': 8,   # Year
            'D': 50,  # Title
            'E': 20,  # Journal
            'F': 30,  # Main idea
            'G': 30,  # Conclusion
            'H': 25,  # Notes 1
            'I': 25,  # Notes 2
            'J': 15,  # Cross-ref
            'K': 30,  # Excerpts
            'L': 40,  # URL
            'M': 40,  # Filename
            'N': 12,  # Downloaded
            'O': 50,  # File Path
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze top row
        ws.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(self.output_excel)
        print(f"Spreadsheet saved: {self.output_excel}")
    
    def generate_summary_report(self):
        """Generate a text summary of the papers found."""
        print("\n" + "="*80)
        print("SUMMARY REPORT")
        print("="*80)
        
        # Count by concept
        concept_counts = {}
        for paper in self.papers:
            concept = paper.concept or 'Uncategorized'
            concept_counts[concept] = concept_counts.get(concept, 0) + 1
        
        print(f"\nPapers by Concept/Topic:")
        for concept, count in sorted(concept_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"  {concept}: {count}")
        
        # Count by year
        year_counts = {}
        for paper in self.papers:
            year = paper.year or 'Unknown'
            year_counts[year] = year_counts.get(year, 0) + 1
        
        print(f"\nPapers by Year:")
        for year, count in sorted(year_counts.items()):
            print(f"  {year}: {count}")
        
        # Download status
        downloaded = sum(1 for p in self.papers if p.is_downloaded)
        print(f"\nDownload Status:")
        print(f"  Downloaded: {downloaded}")
        print(f"  Not downloaded: {len(self.papers) - downloaded}")
        
        # Papers without metadata
        no_title = sum(1 for p in self.papers if not p.title)
        no_author = sum(1 for p in self.papers if not p.author)
        no_year = sum(1 for p in self.papers if not p.year)
        
        print(f"\nMetadata Coverage:")
        print(f"  Missing title: {no_title}/{len(self.papers)}")
        print(f"  Missing author: {no_author}/{len(self.papers)}")
        print(f"  Missing year: {no_year}/{len(self.papers)}")


def main():
    parser = argparse.ArgumentParser(
        description='Organize research papers and generate literature review spreadsheet',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Scan current directory
  python organize_papers.py .
  
  # Scan specific Dropbox folder
  python organize_papers.py ~/Dropbox/Research_Papers
  
  # Specify output Excel file
  python organize_papers.py ~/Dropbox/Research_Papers -o my_literature_review.xlsx
  
  # Include multiple file types
  python organize_papers.py ~/Papers -e .pdf .docx
        """
    )
    
    parser.add_argument('folder', type=str,
                       help='Root folder containing research papers')
    parser.add_argument('-o', '--output', type=str,
                       help='Output Excel filename (default: literature_review_YYYYMMDD.xlsx)')
    parser.add_argument('-e', '--extensions', nargs='+', default=['.pdf'],
                       help='File extensions to scan (default: .pdf)')
    
    args = parser.parse_args()
    
    # Validate folder
    root_path = Path(args.folder).expanduser().resolve()
    if not root_path.exists():
        print(f"Error: Folder does not exist: {root_path}")
        sys.exit(1)
    
    # Create organizer
    organizer = PaperOrganizer(root_path, args.output)
    
    # Scan papers
    organizer.scan_papers(extensions=args.extensions)
    
    # Generate outputs
    if organizer.papers:
        organizer.generate_spreadsheet()
        organizer.generate_summary_report()
    else:
        print("No papers found!")
        sys.exit(1)
    
    print("\n" + "="*80)
    print("Done! Next steps:")
    print("  1. Open the Excel file and review the auto-extracted metadata")
    print("  2. For papers not downloaded, download them from Dropbox")
    print("  3. Re-run this script to update metadata for newly downloaded papers")
    print("  4. Fill in the Main idea, Conclusion, Notes, etc. columns manually")
    print("="*80)


if __name__ == '__main__':
    main()
